#!/usr/bin/env python3
"""
CRM Lead Entry Automation for Fasttrack CRM

Usage:
    python3 crm_entry.py                  # Process all 15 trial leads (rows 132-146)
    python3 crm_entry.py --dry-run        # Show what would be done
    python3 crm_entry.py --row 139        # Process a single row
    python3 crm_entry.py --rows 132-135   # Process a range

Process per lead:
  1. Search %LastName% in sidebar
  2. If person found at same institution:
       Click contact -> New Note (paste text) -> Save -> New Call (Email intro, Low) -> Save
  3. If not found but company exists:
       Click org -> Add contact (first, last, email, job title) -> then do notes/call
  4. If company doesn't exist:
       New Company -> fill org name + "xxx" switchboard -> add contact -> notes/call
"""

import os, sys, time, re, json, argparse
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# ============ CONFIG ============
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "APA 2026 (1).xlsx")
SHEET_NAME = "Ben Leads"
CRM_URL = "https://websedge.axlr8.uk"
CRM_USER = "Ben Varvill"
CRM_PASS = "Jrm4AEZa"
DATA_DIR = "/tmp/chrome_crm_auto"

COL_INSTITUTION = 4
COL_POSITION = 7
COL_CONTACT_NAME = 8
COL_EMAIL = 9
COL_PASTE_CRM = 34
DEFAULT_ROWS = (132, 146)

WAIT = 3  # seconds between actions


def parse_results(html):
    """Parse search results HTML. Returns list of dicts with org_name, org_key, contact_name, contact_key."""
    results = []
    for m in re.finditer(
        r"<tr[^>]*id=\"res(\d+)\"[^>]*>(.*?)</tr>", html, re.DOTALL
    ):
        row = {"org_name": "", "org_key": "", "contact_name": "", "contact_key": ""}
        for link in re.finditer(
            r"cOn\('\d+','([cos])(\d+)','b'\)[^>]*>([^<]+)", m.group(2)
        ):
            typ, key, text = link.group(1), link.group(2), link.group(3).strip()
            text = re.sub(r'\s+', ' ', text)
            if typ == 'o':
                row["org_name"] = text
                row["org_key"] = key
            elif typ == 'c':
                row["contact_name"] = text
                row["contact_key"] = key
        if row["contact_name"] or row["org_name"]:
            results.append(row)
    return results


class CRM:
    def __init__(self, dry_run=False):
        self.dry = dry_run
        self.d = None  # selenium driver
        self.main = None  # main window handle

    def login(self):
        print("Starting Chrome + logging in...")
        opts = webdriver.ChromeOptions()
        opts.add_argument(f"--user-data-dir={DATA_DIR}")
        opts.add_argument("--disable-popup-blocking")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        self.d = webdriver.Chrome(options=opts)
        self.d.set_window_size(1400, 900)
        self.d.get(f"{CRM_URL}/login.asp")
        time.sleep(2)
        if "login" in self.d.current_url.lower():
            self.d.find_element(By.NAME, "username").send_keys(CRM_USER)
            self.d.find_element(By.NAME, "userpass").send_keys(CRM_PASS + Keys.RETURN)
            time.sleep(5)
        assert "Fasttrack" in self.d.title, f"Login failed: {self.d.title}"
        self.main = self.d.current_window_handle
        print(f"  Logged in: {self.d.title}\n")

    def _default(self):
        self.d.switch_to.default_content()

    def _close_popups(self):
        for h in self.d.window_handles:
            if h != self.main:
                try:
                    self.d.switch_to.window(h)
                    self.d.close()
                except: pass
        self.d.switch_to.window(self.main)
        self._default()

    # ---- SEARCH ----

    def search_lastname(self, lastname):
        """Search by last name. Returns parsed results."""
        self._default()
        self.d.execute_script("il_show_search()")
        time.sleep(0.5)
        box = self.d.find_element(By.ID, "lms_lastname_search")
        box.clear()
        box.send_keys(f"%{lastname}%" + Keys.RETURN)
        time.sleep(WAIT)
        self.d.switch_to.frame("search_results")
        html = self.d.execute_script("return document.body.innerHTML")
        self._default()
        return parse_results(html)

    def search_org(self, orgname):
        """Search by org name. Returns parsed results."""
        self._default()
        self.d.execute_script("il_show_search()")
        time.sleep(0.5)
        # Clear lastname box first
        try:
            self.d.find_element(By.ID, "lms_lastname_search").clear()
        except: pass
        box = self.d.find_element(By.ID, "lms_orgname_search")
        box.clear()
        box.send_keys(f"%{orgname}%" + Keys.RETURN)
        time.sleep(WAIT)
        self.d.switch_to.frame("search_results")
        html = self.d.execute_script("return document.body.innerHTML")
        self._default()
        return parse_results(html)

    def find_contact(self, results, firstname, lastname, institution):
        """Find a contact matching name + institution in search results."""
        fn = firstname.lower()
        ln = lastname.lower()
        inst_words = [w.lower() for w in institution.split() if len(w) > 3]

        for r in results:
            c = r["contact_name"].lower()
            o = r["org_name"].lower()
            # Must match lastname
            if ln not in c:
                continue
            # Check firstname
            if fn and fn not in c:
                continue
            # Check institution
            if inst_words and any(w in o for w in inst_words):
                return r
            elif institution.lower() in o:
                return r
        return None

    def find_org(self, results, institution):
        """Find an org matching institution in results."""
        inst_words = [w.lower() for w in institution.split() if len(w) > 3]
        for r in results:
            o = r["org_name"].lower()
            if inst_words and any(w in o for w in inst_words):
                return r
            elif institution.lower() in o:
                return r
        return None

    # ---- LOAD CONTACT & NOTES TAB ----

    def load_contact(self, contact_key):
        """Load a contact's details + notes tab. contact_key is the number."""
        ak = f"con{contact_key}"
        self._default()
        self.d.execute_script(f"doPopOverviewFromQs('?actionkey={ak}')")
        time.sleep(2)
        self.d.execute_script(f"il_show_notes('{ak}')")
        time.sleep(2)
        return ak

    # ---- NEW NOTE (inline in detailsframe2) ----

    def create_note(self, contact_key, note_text):
        """Click New Note in detailsframe2, type text, save."""
        ak = f"con{contact_key}"
        if self.dry:
            print(f"    [DRY] Note for {ak}: {note_text[:80]}...")
            return True

        self._default()
        self.d.switch_to.frame("detailsframe2")

        # Click "New Note for ..."
        clicked = False
        for lnk in self.d.find_elements(By.TAG_NAME, "a"):
            if "New Note" in (lnk.text or ""):
                lnk.click()
                clicked = True
                break
        if not clicked:
            print("    ERROR: 'New Note' link not found")
            self._default()
            return False

        time.sleep(2)

        # Fill the textarea (use JS because it may be behind a rich text editor)
        try:
            self.d.execute_script(
                "document.notes_form.new_Note_text.value = arguments[0]",
                note_text
            )
        except Exception as e:
            # Fallback: try direct interaction
            try:
                ta = self.d.find_element(By.NAME, "new_Note_text")
                self.d.execute_script("arguments[0].style.display='block'; arguments[0].focus();", ta)
                ta.clear()
                ta.send_keys(note_text)
            except Exception as e2:
                print(f"    ERROR filling note: {e2}")
                self._default()
                return False

        time.sleep(0.5)

        # Click "Save Note for ..."
        saved = False
        for lnk in self.d.find_elements(By.TAG_NAME, "a"):
            if "Save Note" in (lnk.text or ""):
                lnk.click()
                saved = True
                break
        if not saved:
            # Fallback: JS save
            try:
                self.d.execute_script("doRichTextSaveNote(document.notes_form)")
            except:
                print("    ERROR: Could not save note")
                self._default()
                return False

        time.sleep(WAIT)
        self._default()
        print("    Note saved")
        return True

    # ---- NEW CALL (popup) ----

    def create_call(self, contact_key):
        """Click New Call in detailsframe2 -> popup -> fill -> save."""
        if self.dry:
            print(f"    [DRY] Call for con{contact_key}: Email intro, Low")
            return True

        self._default()
        self.d.switch_to.frame("detailsframe2")

        # Click "New Call"
        handles_before = set(self.d.window_handles)
        clicked = False
        for lnk in self.d.find_elements(By.TAG_NAME, "a"):
            if lnk.text.strip() == "New Call":
                lnk.click()
                clicked = True
                break
        self._default()

        if not clicked:
            print("    ERROR: 'New Call' link not found")
            return False

        time.sleep(WAIT)

        # Switch to popup
        new_handles = set(self.d.window_handles) - handles_before
        if not new_handles:
            print("    ERROR: Call popup didn't open")
            return False

        self.d.switch_to.window(list(new_handles)[0])

        try:
            # Set "Call For" to Ben Varvill
            self.d.execute_script("""
                var s = document.callDetails.createdFor;
                for (var i=0; i<s.options.length; i++)
                    if (s.options[i].text.indexOf('Ben Varvill') >= 0) { s.selectedIndex=i; break; }
            """)

            # Set "Type of Call" to Intro Email
            self.d.execute_script("""
                var s = document.callDetails.typeOfCall;
                for (var i=0; i<s.options.length; i++)
                    if (s.options[i].text.indexOf('Intro Email') >= 0) { s.selectedIndex=i; break; }
            """)

            # Set call reason
            self.d.execute_script("document.callDetails.callReason.value = 'Email intro'")

            # Set priority to Low (radio value='L')
            self.d.execute_script("""
                var r = document.callDetails.callPriority;
                for (var i=0; i<r.length; i++)
                    if (r[i].value === 'L') { r[i].checked = true; break; }
            """)

            # Click Save
            for lnk in self.d.find_elements(By.TAG_NAME, "a"):
                if lnk.text.strip() == "Save":
                    lnk.click()
                    break

            time.sleep(WAIT)
        except Exception as e:
            print(f"    ERROR filling call: {e}")

        # Close popup, back to main
        try:
            self.d.close()
        except: pass
        self.d.switch_to.window(self.main)
        self._default()

        print("    Call saved (Low, Email intro)")
        return True

    # ---- NEW COMPANY + CONTACT ----

    def create_company_and_contact(self, org_name, first_name, last_name, email, job_title):
        """Click 'New Company', fill org + contact details, save."""
        if self.dry:
            print(f"    [DRY] New company: {org_name}")
            print(f"    [DRY] New contact: {first_name} {last_name}, {email}")
            return True

        self._default()
        # Click "New Company" in left sidebar
        self.d.execute_script("il_show_newosc()")
        time.sleep(2)

        # Fill in oscframe
        self.d.switch_to.frame("oscframe")

        try:
            # Org name
            self.d.find_element(By.NAME, "org_name").clear()
            self.d.find_element(By.NAME, "org_name").send_keys(org_name)

            # Switchboard = "xxx"
            try:
                self.d.find_element(By.NAME, "site_telephone").clear()
                self.d.find_element(By.NAME, "site_telephone").send_keys("xxx")
            except: pass

            # Contact details
            self.d.find_element(By.NAME, "cont_1stname").clear()
            self.d.find_element(By.NAME, "cont_1stname").send_keys(first_name)
            self.d.find_element(By.NAME, "cont_lastname").clear()
            self.d.find_element(By.NAME, "cont_lastname").send_keys(last_name)
            self.d.find_element(By.NAME, "cont_email").clear()
            self.d.find_element(By.NAME, "cont_email").send_keys(email)
            self.d.find_element(By.NAME, "cont_jobtitle").clear()
            self.d.find_element(By.NAME, "cont_jobtitle").send_keys(job_title)

            # Save all-in-one form
            self.d.execute_script("document.b2bAllIn1.submit()")
            time.sleep(WAIT)
            print(f"    Company + contact created")
        except Exception as e:
            print(f"    ERROR creating company/contact: {e}")
            self._default()
            return False

        self._default()
        return True

    def add_contact_to_org(self, org_key, first_name, last_name, email, job_title):
        """Load an existing org, then add a contact."""
        if self.dry:
            print(f"    [DRY] Add contact to org{org_key}: {first_name} {last_name}")
            return True

        self._default()
        # Load the org
        self.d.execute_script(f"doPopOverviewFromQs('?actionkey=org{org_key}')")
        time.sleep(WAIT)

        # Fill contact in oscframe
        self.d.switch_to.frame("oscframe")
        try:
            self.d.find_element(By.NAME, "cont_1stname").clear()
            self.d.find_element(By.NAME, "cont_1stname").send_keys(first_name)
            self.d.find_element(By.NAME, "cont_lastname").clear()
            self.d.find_element(By.NAME, "cont_lastname").send_keys(last_name)
            self.d.find_element(By.NAME, "cont_email").clear()
            self.d.find_element(By.NAME, "cont_email").send_keys(email)
            self.d.find_element(By.NAME, "cont_jobtitle").clear()
            self.d.find_element(By.NAME, "cont_jobtitle").send_keys(job_title)

            # Save contact form
            self.d.execute_script("document.contDetails.submit()")
            time.sleep(WAIT)
            print(f"    Contact added to org")
        except Exception as e:
            print(f"    ERROR adding contact: {e}")
            self._default()
            return False

        self._default()
        return True

    # ---- MAIN PROCESS ----

    def process_lead(self, lead):
        name = lead["name"]
        inst = lead["institution"]
        pos = lead["position"]
        email = lead["email"]
        paste = lead["paste_crm"]

        parts = name.split()
        first = parts[0]
        last = " ".join(parts[1:]) if len(parts) > 1 else parts[0]

        print(f"\n{'='*50}")
        print(f"  {name} | {inst}")
        print(f"  {pos} | {email}")
        print(f"{'='*50}")

        if not paste:
            print("  SKIP: No paste text")
            return False

        # Step 1: Search by last name
        print(f"  Searching '{last}'...")
        results = self.search_lastname(last)
        print(f"  {len(results)} results")

        match = self.find_contact(results, first, last, inst)

        if match:
            # SCENARIO 1: Contact found at institution
            ck = match["contact_key"]
            print(f"  FOUND: {match['contact_name']} at {match['org_name']} (con{ck})")

            self.load_contact(ck)
            self.create_note(ck, paste)
            # Reload notes tab so "New Call" link is visible again
            self.load_contact(ck)
            self.create_call(ck)
            return True

        # Step 2: Search for company
        print(f"  Contact not found. Searching org '{inst}'...")
        org_results = self.search_org(inst)
        print(f"  {len(org_results)} org results")

        org_match = self.find_org(org_results, inst)
        job_title = f"{pos} at {inst}" if pos else inst

        if org_match:
            # SCENARIO 2: Org exists, contact doesn't
            ok = org_match["org_key"]
            print(f"  FOUND ORG: {org_match['org_name']} (org{ok})")
            print(f"  Adding contact...")
            self.add_contact_to_org(ok, first, last, email, job_title)

            # Re-search to find the new contact
            time.sleep(1)
            results2 = self.search_lastname(last)
            match2 = self.find_contact(results2, first, last, inst)
            if match2:
                ck = match2["contact_key"]
                print(f"  New contact: con{ck}")
                self.load_contact(ck)
                self.create_note(ck, paste)
                self.load_contact(ck)
                self.create_call(ck)
                return True
            else:
                print("  ERROR: Can't find newly added contact")
                return False
        else:
            # SCENARIO 3: Neither exists
            print(f"  Org not found. Creating new company + contact...")
            self.create_company_and_contact(inst, first, last, email, job_title)

            # Re-search
            time.sleep(1)
            results3 = self.search_lastname(last)
            match3 = self.find_contact(results3, first, last, inst)
            if match3:
                ck = match3["contact_key"]
                print(f"  New contact: con{ck}")
                self.load_contact(ck)
                self.create_note(ck, paste)
                self.load_contact(ck)
                self.create_call(ck)
                return True
            else:
                print("  ERROR: Can't find newly created contact")
                return False


def load_leads(start, end):
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    leads = []
    for r in range(start, end + 1):
        name = (ws.cell(r, COL_CONTACT_NAME).value or "").strip()
        if not name:
            continue
        email = (ws.cell(r, COL_EMAIL).value or "").strip()
        if email.startswith("2x"):
            email = email[2:].strip().split(";")[0].strip()
        leads.append({
            "row": r,
            "name": name,
            "institution": (ws.cell(r, COL_INSTITUTION).value or "").strip(),
            "position": (ws.cell(r, COL_POSITION).value or "").strip(),
            "email": email,
            "paste_crm": (str(ws.cell(r, COL_PASTE_CRM).value or "")).strip(),
        })
    wb.close()
    return leads


def load_leads_from_file(filepath):
    """Load leads from a pipe-delimited text file.
    Format per line: Institution | Position | Contact Name | Email | Paste in CRM text
    Lines starting with # are ignored. Use \\n for newlines within the paste text.
    """
    leads = []
    with open(filepath, 'r') as f:
        for i, line in enumerate(f, 1):
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            parts = [p.strip() for p in line.split('|')]
            if len(parts) < 5:
                print(f"  WARNING: Line {i} has {len(parts)} fields (need 5), skipping")
                continue
            institution, position, name, email, paste = parts[0], parts[1], parts[2], parts[3], '|'.join(parts[4:])
            # Convert literal \n to actual newlines in paste text
            paste = paste.replace('\\n', '\n')
            # Clean email
            if email.startswith("2x"):
                email = email[2:].strip().split(";")[0].strip()
            leads.append({
                "row": i,
                "name": name.strip(),
                "institution": institution.strip(),
                "position": position.strip(),
                "email": email.strip(),
                "paste_crm": paste.strip(),
            })
    return leads


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--row", type=int)
    ap.add_argument("--rows", type=str)
    ap.add_argument("--file", type=str,
                    help="Load leads from a pipe-delimited text file instead of Excel")
    args = ap.parse_args()

    if args.file:
        leads = load_leads_from_file(args.file)
        print(f"Loaded {len(leads)} leads from {args.file}:\n")
        for l in leads:
            print(f"  {l['name']} at {l['institution']}")
    else:
        s, e = DEFAULT_ROWS
        if args.row:
            s = e = args.row
        elif args.rows:
            parts = args.rows.split("-")
            s = int(parts[0])
            e = int(parts[1]) if len(parts) > 1 else s

        leads = load_leads(s, e)
        print(f"Loaded {len(leads)} leads (rows {s}-{e}):\n")
        for l in leads:
            print(f"  Row {l['row']}: {l['name']} at {l['institution']}")

    if not leads:
        print("Nothing to do.")
        return

    crm = CRM(dry_run=args.dry_run)

    if not args.dry_run:
        print(f"\n  About to process {len(leads)} leads in the CRM.")
        resp = input("  Type YES to proceed: ")
        if resp.strip() != "YES":
            print("Aborted.")
            return

    crm.login()

    ok = fail = 0
    for lead in leads:
        try:
            if crm.process_lead(lead):
                ok += 1
            else:
                fail += 1
        except Exception as ex:
            print(f"  EXCEPTION: {ex}")
            import traceback; traceback.print_exc()
            fail += 1
            try: crm._close_popups()
            except: pass
        time.sleep(1)

    print(f"\n{'='*50}")
    print(f"Done: {ok} succeeded, {fail} failed")
    print(f"{'='*50}")

    if not args.dry_run:
        try:
            input("Press Enter to close browser...")
        except EOFError:
            pass
        crm.d.quit()


if __name__ == "__main__":
    main()
